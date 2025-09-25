from flask import Flask, request, jsonify
import requests
import os
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
import shutil
from win32com import client
import pythoncom
import datetime

load_dotenv()

MONDAY_API_URL = "https://api.monday.com/v2"
MONDAY_FILE_URL = "https://api.monday.com/v2/file"

API_KEY = os.getenv("MONDAY_API_KEY")
BOARD_ID = os.getenv("MONDAY_CCINVOICE_BOARD_ID")
GROUP_ID = "topics"

empty_PPA = "CCPPA Form.xlsx"

column_ids = {'Vendor':'text_mkv6s9er', 'Statement':'color_mkvkk3zr','Total':'numeric_mkv65026','Inkind':'numeric_mkv6tbbk','Department':'dropdown_mkvkjm3z','Description':'text_mkve5mct','OrderedBy':'multiple_person_mkvkp06f','PPA file':'file_mkvksbn1','Workflow':'color_mkv67wpq','status':'status'}


app = Flask(__name__)

HEADERS = {
    "Authorization" : API_KEY,
}


class add_CCPPA_to_board:

    def __init__(self, board_id, API_KEY):
        self.board_id = board_id
        self.API_KEY = API_KEY
        self.current_invoice_number = 0
        self.group_ids = []
        self.current_item_id = 0
        self.current_item_data = {}
        self.group_statement = ''
        return
    
    def upload_to_monday(self, file_path):
        #create query and send file
        item_id = self.current_item_id
        column_id = column_ids['PPA file']
        file = file_path
        if not item_id or not column_id or not file:
            return jsonify({"error": "Missing item_id, column_id, or file"}), 400
        query = f'''
        mutation ($file: File!) {{
        add_file_to_column (file: $file, item_id: {item_id}, column_id: "{column_id}") {{
            id
        }}
        }}
        '''
        #open file to prepare to send
        PPA_FILE = open(file_path,'rb')
        multipart_data = {
            'query': (None, query),
            'variables[file]': PPA_FILE
        }
        response = requests.post(MONDAY_FILE_URL, headers=HEADERS, files=multipart_data)
        #close file
        PPA_FILE.close()
        #check for errors
        if response.status_code != 200:
            print(f"HTTP Error while uploading: {response.status_code}")
            print(response.text)
        else:
            response_json = response.json()
            # Check for GraphQL errors inside the response body
            if 'errors' in response_json:
                print("GraphQL Errors found during file upload:")
                for error in response_json['errors']:
                    print(f"- Message: {error.get('message')}")
                    if 'locations' in error:
                        print(f"  Location: {error['locations']}")
                    if 'path' in error:
                        print(f"  Path: {error['path']}")
            else:
                filetype = file_path[file_path.find('.')+1:]
                print(filetype + 'File Uploaded')
        try:
            os.remove(file_path)
            print(f"File '{file_path}' deleted successfully.")
        except Exception as e:
            print(f"An error occurred while deleting" + filetype + "file: {e}")
        return
    
    def get_item(self, item_id):
        self.current_item_id = item_id
        data_good = False
        self.current_item_data = {}
        # Query Monday Board for Item info
        query = {
            'query': f'''
                query {{
                    items(ids: {item_id}) {{
                        id
                        name
                        column_values {{
                            id
                            column{{
                                title
                            }}
                            text
                        }}
                    }}
                }}
            '''
        }
        response = requests.post(MONDAY_API_URL, json=query, headers=HEADERS)
        #parse response for column values and check for errors
        if response.status_code != 200:
            print(f"HTTP Error: {response.status_code}")
            print(response.text)
        else:
            response_json = response.json()
            # Check for GraphQL errors inside the response body
            if 'errors' in response_json:
                print("GraphQL Errors found:")
                for error in response_json['errors']:
                    print(f"- Message: {error.get('message')}")
                    if 'locations' in error:
                        print(f"  Location: {error['locations']}")
                    if 'path' in error:
                        print(f"  Path: {error['path']}")
            else:
                # No GraphQL errors; process the data
                item_data = response_json['data']['items'][0]
                self.current_invoice_number = item_data['name']
                #print(f"Invoice #: {item_data['name']}")
                #print("Column Values:")
                # loop through column_values and extract data needed
                for column in item_data['column_values']:
                    self.current_item_data[column['id']] = (column['text'])
                    #print(f"- {column['column']['title']} ({column['id']}): {column['text']}")
                #print(self.current_item_data[column_ids['PPA file']])
                if not self.current_item_data[column_ids['PPA file']]:
                    if self.current_item_data[column_ids['Workflow']] == 'PPA Creation':
                        print("Data is good to go")
                        data_good = True
        return data_good
    
    def cc_get_item(self, item_id):
        self.current_item_id = item_id
        data_good = False
        self.current_item_data = {}
        # Query Monday Board for Item info
        query = {
            'query': f'''
                query {{
                    items(ids: {item_id}) {{
                        id
                        name
                        column_values {{
                            id
                            column{{
                                title
                            }}
                            text
                        }}
                    }}
                }}
            '''
        }
        response = requests.post(MONDAY_API_URL, json=query, headers=HEADERS)
        #parse response for column values and check for errors
        if response.status_code != 200:
            print(f"HTTP Error: {response.status_code}")
            print(response.text)
        else:
            response_json = response.json()
            # Check for GraphQL errors inside the response body
            if 'errors' in response_json:
                print("GraphQL Errors found:")
                for error in response_json['errors']:
                    print(f"- Message: {error.get('message')}")
                    if 'locations' in error:
                        print(f"  Location: {error['locations']}")
                    if 'path' in error:
                        print(f"  Path: {error['path']}")
            else:
                # No GraphQL errors; process the data
                item_data = response_json['data']['items'][0]
                self.current_invoice_number = item_data['name']
                #print(f"Invoice #: {item_data['name']}")
                #print("Column Values:")
                # loop through column_values and extract data needed
                for column in item_data['column_values']:
                    self.current_item_data[column['id']] = (column['text'])
                    #print(f"- {column['column']['title']} ({column['id']}): {column['text']}")
                #print(self.current_item_data[column_ids['PPA file']])
                if self.current_item_data[column_ids['Workflow']] == 'Credit Card PPA':
                    #print("Data is good to go")
                    data_good = True
        return data_good
    
    def xlsxtopdf(self, file_path):
        #initialize pythoncom
        excel = client.Dispatch("Excel.Application",pythoncom.CoInitialize())
        #open workbook
        sheets = excel.Workbooks.Open(os.path.abspath(file_path))
        work_sheets = sheets.Worksheets[0]
        #export as pdf
        work_sheets.ExportAsFixedFormat(0, os.path.abspath(file_path[:len(file_path)-4] + 'pdf'))
        #close workbook and exit pythoncom
        sheets.Close()
        excel.Quit()
        print('PDF created successfully')
        try:
            os.remove(file_path)
            print(f"File '{file_path}' deleted successfully.")
        except Exception as e:
            print(f"An error occurred while deleting xlsx file: {e}")
        return

    def ccxlsxtopdf(self, file_path):
        #initialize pythoncom
        excel = client.Dispatch("Excel.Application",pythoncom.CoInitialize())
        #open workbook
        sheets = excel.Workbooks.Open(os.path.abspath(file_path))
        work_sheets = sheets.Worksheets[0]
        #export as pdf
        work_sheets.ExportAsFixedFormat(0, os.path.abspath(file_path[:len(file_path)-4] + 'pdf'))
        #close workbook and exit pythoncom
        sheets.Close()
        excel.Quit()
        print('PDF created successfully')
        return

    def create_ppa(self):
        data = self.current_item_id
        new_ppa_name ="PPA Form -" + data[column_ids['Vendor']] + self.current_invoice_number + ".xlsx"
        #destination_file = destination_folder + r"/" + new_ppa_name
        destination_file = new_ppa_name
        #Create the file for ppa
        try:
            shutil.copy(empty_PPA, destination_file)
            print(f"'{empty_PPA}' copied to '{destination_file}' successfully.")
        except FileNotFoundError:
            print(f"Error: '{empty_PPA}' not found.")
        except Exception as e:
            print(f"An error occurred when creating empty PPA form: {e}")
        #open file
        try:
            workbook = load_workbook(destination_file)
            sheet = workbook.active
            sheet["L13"] = data[column_ids['Vendor']]
            sheet["B13"] = data[column_ids['Department']]
            sheet["G19"] = self.current_invoice_number
            sheet["B23"] = 1
            sheet["N23"] = data[column_ids['Total']]
            sheet["C23"] = data[column_ids['Description']]
            sheet["C24"] = "job date: " + data[column_ids['Date']]
            sheet["D40"] = "PPA Prepared by " + data[column_ids['OrderedBy']]
            workbook.save(filename=destination_file)
        except Exception as e:
            print(e)
        else:
            print('PPA Created Succesfully')
        self.xlsxtopdf(destination_file)
        self.upload_to_monday(destination_file[:len(destination_file)-4]+'pdf')
        return True

    def get_group_ids(self, group_id):
        item_ids = []
        query = """
        query ($board_id: ID!, $group_id: String!) {
        boards(ids: [$board_id]) {
            groups(ids: [$group_id]) {
                id
                title
                items_page{
                items {
                    id
                    name
                }
                }
            }
        }
        }
        """
        variables = {
            "board_id" : BOARD_ID,
            "group_id" : group_id
        }
        full_query = {"query": query, "variables": variables}
        response = requests.post(MONDAY_API_URL, json=full_query, headers=HEADERS)
        # Check for GraphQL errors inside the response body
        if response.status_code != 200:
                print(f"HTTP Error getting item ids: {response.status_code}")
                print(response.text)
        else:
            response_json = response.json()
            # Check for GraphQL errors inside the response body
            if 'errors' in response_json:
                print("GraphQL Errors found getting item ids:")
                for error in response_json['errors']:
                    print(f"- Message: {error.get('message')}")
                    if 'locations' in error:
                        print(f"  Location: {error['locations']}")
                    if 'path' in error:
                        print(f"  Path: {error['path']}")
            else:
                items = response_json['data']['boards'][0]['groups'][0]['items_page']['items']
                #print("Error:", response.text)
                for item in items:
                    item_ids.append(item['id'])
                #print("Item IDs:", item_ids)
        self.group_ids = item_ids
        return item_ids
    
    def create_ccppa(self, group):
        items = self.get_group_ids(group)
        nancy_items = []
        tyson_items = []
        add.cc_get_item(self.group_ids[0])
        self.group_statement = self.current_item_data[column_ids['Statement']]
        if not self.group_statement:
            print("Error retrieving statement. Exiting")
            return
        for item in items:
            #print('Started PPA creation on item: ' + str(item))
            good_data = add.cc_get_item(item)
            if good_data == True:
                ordered_by = self.current_item_data[column_ids['OrderedBy']]
                if ordered_by == "Nancy Ortega":
                    nancy_items.append(self.current_item_data)
                elif ordered_by == "Tyson Murphy":
                    tyson_items.append(self.current_item_data)
                else:
                    print("error: an item is missing who it was purchased by. Excluding item")
            else:
                print('Item # ' + str(self.current_item_id) + ' is empty or there was an error')

        new_ppa_name = self.group_statement + "Elan Financial Serv PPA.xlsx"
        #destination_file = destination_folder + r"/" + new_ppa_name
        #currently destination folder same as working directory
        destination_file = new_ppa_name
        #Create the file for ppa
        try:
            shutil.copy(empty_PPA, destination_file)
            print(f"'{empty_PPA}' copied to '{destination_file}' successfully.")
        except FileNotFoundError:
            print(f"Error: '{empty_PPA}' not found.")
        except Exception as e:
            print(f"An error occurred when creating empty PPA form: {e}")
        #open file
        try:
            workbook = load_workbook(destination_file)
            total = 0
            sheet = workbook.active
            current_row = 24
            sheet["C23"] = "Statement for Nancy Ortega"
            current_row = 24
            for item in nancy_items:
                sheet["B"+str(current_row)] = "1"
                sheet["C"+str(current_row)] = item[column_ids['Description']]
                sheet["M"+str(current_row)] = item[column_ids['Department']]
                sheet["N"+str(current_row)] = item[column_ids['Total']]
                total += float(item[column_ids['Total']])
                current_row += 1
            
            current_row += 1
            sheet["C"+str(current_row)] = "Statement for Tyson Murphy"
            current_row += 1
            for item in tyson_items:
                sheet["B"+str(current_row)] = "1"
                sheet["C"+str(current_row)] = item[column_ids['Description']]
                sheet["M"+str(current_row)] = item[column_ids['Department']]
                sheet["N"+str(current_row)] = item[column_ids['Total']]
                total += float(item[column_ids['Total']])
                current_row += 1
            workbook.save(filename=destination_file)
        except Exception as e:
            print(e)
        else:
            print('PPA Created Succesfully')
        self.ccxlsxtopdf(destination_file)
        #create new item to hold PPA
        today_date = datetime.datetime.now().strftime("%Y-%m-%d")
        print(str(today_date))

        column_values = '''{
            "numeric_mkv6v1a1": '''+ str(total) +''',  
            "color_mkv67wpq": "PPA Creation",
            "status" : "Waiting for Tyson to Sign PPA",
            "person" : "nancy.ortega@ochsinc.org",
            "date_mkvpfs6q" : "'''+ today_date +'''",
            "dropdown_mkvkjm3z" : "Various"
        }'''
        query = """
            mutation ($boardId: ID!, $groupId: String!, $itemName: String!, $columnVals: JSON!){
                create_item (
                    board_id: $boardId,
                    group_id: $groupId,
                    item_name: $itemName,
                    column_values: $columnVals
                ) {
                    id
                }
            }
        """
        variables = {
            "boardId" : BOARD_ID,
            "groupId" : GROUP_ID,
            "itemName": "PPA - " + self.group_statement,
            "columnVals" : column_values
        }
        data = {
            "query": query,
            "variables": variables
        }
        response = requests.post(MONDAY_API_URL, json=data, headers=HEADERS)
        #check for errors in new item
        if response.status_code != 200:
            print(f"Error creating new item: {response.status_code}")
            print(response.text)
        else:
            response_json = response.json()
            # Check for GraphQL errors inside the response body
            if 'errors' in response_json:
                print("GraphQL Errors found during item creation:")
                for error in response_json['errors']:
                    print(f"- Message: {error.get('message')}")
                    if 'locations' in error:
                        print(f"  Location: {error['locations']}")
                    if 'path' in error:
                        print(f"  Path: {error['path']}")
                #delete pdf and return
                try:
                    os.remove(destination_file)
                    print(f"File '{destination_file}' deleted successfully.")
                except Exception as e:
                    print(f"An error occurred while deleting PDF file: {e}")
                return
            else:
                print('item created')
                self.current_item_id = response_json['data']['create_item']['id']
                self.upload_to_monday(destination_file[:len(destination_file)-4]+'pdf')
                self.upload_to_monday(destination_file)
        
        return True

add = add_CCPPA_to_board(BOARD_ID, API_KEY)

#need to redo
@app.route("/add_ccppa", methods=["POST"])
def add_ccppa():
    data = request.json
    if data:
        print("Request recieved\n")
    if not data:
        print("Request empty")
        return ":("
    group_id = data["event"]["groupId"]
    if group_id:
        add.create_ccppa(group_id)
    else:
        print('Stopped Creation process, no group presented')
    return 'done'


if __name__ == "__main__":
    app.run(debug=True)
    #add.create_ccppa(GROUP_ID)